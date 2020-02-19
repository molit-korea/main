# -*- coding: utf-8 -*-
"""
Spyder Editor

This is a temporary script file.
"""

import networkx as nx
import matplotlib.pyplot as plt
from openpyxl import load_workbook
import os


file_path = os.path.dirname(os.path.realpath(__file__))

load_wb_school = load_workbook("{}/data/school_zone_accident.xlsx".format(file_path), data_only=True)
load_ws_school = load_wb_school['Sheet1']
load_wb_adult = load_workbook("{}/data/adult_zone_accident.xlsx".format(file_path), data_only=True)
load_ws_adult = load_wb_adult['Sheet1']


def draw_graph(input_G, layout="shell"):
    if layout=="shell":
        pos = nx.shell_layout(input_G)
    elif layout=="spring":
        pos = nx.spring_layout(input_G)
    elif layout=="spectral":
        pos = nx.spectral_layout(input_G)
    elif layout=="circular":
        pos = nx.circular_layout(input_G)
    elif layout=="random":
        pos = nx.random_layout(input_G)
    else:
        pos = nx.shell_layout(input_G)
        
    plt.figure()
    nx.draw_networkx_nodes(input_G, pos)
    nx.draw_networkx_edges(input_G, pos)
    nx.draw_networkx_labels(input_G, pos)
    nx.draw_networkx_edge_labels(input_G, pos)
    #plt.axis("off")
    


zone_type = ["School Zone", "Old Man Zone"]
accident_type = ["Car", "Personal Mobility"]



MDG = nx.MultiDiGraph()
MDG.add_edges_from([(accident_type[0], zone_type[0]), (accident_type[0], zone_type[1]), (accident_type[1], zone_type[0]), (accident_type[1], zone_type[1])])
MDG.add_edges_from([ (accident_type[0], zone_type[0]) for i in range(0, int(load_ws_school['D6'].value))])
MDG.add_edges_from([ (accident_type[0], zone_type[1]) for i in range(0, int(load_ws_adult['D6'].value))])
MDG.add_edges_from([ (accident_type[1], zone_type[0]) for i in range(0, int(load_ws_school['D5'].value))])
MDG.add_edges_from([ (accident_type[1], zone_type[1]) for i in range(0, int(load_ws_adult['D5'].value))])
draw_graph(MDG)
plt.savefig("{}/final_model.png".format(file_path))




    